#!/usr/bin/env python3
"""Analizza i percentili delle statistiche /90 per ruolo, per calibrare le scale fisse dello score."""
import openpyxl

MASTER = "/home/ubuntu/Downloads/euroleghe_master_classic_fotmob_2026_27_final_clean.xlsx"
wb = openpyxl.load_workbook(MASTER, data_only=True)
ws = wb["Solo_Torneo"]
h = [c.value for c in ws[1]]
c = {v: i + 1 for i, v in enumerate(h) if v}


def n(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def pct(vals, p):
    if not vals:
        return 0.0
    s = sorted(vals)
    idx = min(len(s) - 1, int(round(p / 100.0 * (len(s) - 1))))
    return s[idx]


# raccogli per ruolo
data = {r: [] for r in ("P", "D", "C", "A")}
for r in range(2, ws.max_row + 1):
    ruolo = ws.cell(r, c["ruolo"]).value
    if ruolo not in data:
        continue
    minuti = n(ws.cell(r, c["minuti_campionato"]).value)
    pres = n(ws.cell(r, c["presenze_campionato"]).value)
    if minuti <= 0:
        continue
    rec = {
        "minuti": minuti,
        "pres": pres,
        "gol90": n(ws.cell(r, c["gol_campionato"]).value) / minuti * 90,
        "ass90": n(ws.cell(r, c["assist_campionato"]).value) / minuti * 90,
        "npxG90": n(ws.cell(r, c["xG_no_rigori"]).value) / minuti * 90,
        "xA90": n(ws.cell(r, c["xA"]).value) / minuti * 90,
        "tocchi90": n(ws.cell(r, c["tocchi_area"]).value) / minuti * 90,
        "tiri90": n(ws.cell(r, c["tiri"]).value) / minuti * 90,
        "tiri_porta90": n(ws.cell(r, c["tiri_in_porta"]).value) / minuti * 90,
        "occ90": n(ws.cell(r, c["occasioni_create"]).value) / minuti * 90,
        "cs_rate": (n(ws.cell(r, c["clean_sheet"]).value) / pres) if pres > 0 else 0.0,
        "sub90": n(ws.cell(r, c["gol_subiti"]).value) / minuti * 90,
    }
    data[ruolo].append(rec)

KEYS = {
    "P": ["cs_rate", "sub90"],
    "D": ["gol90", "ass90", "npxG90", "xA90", "tocchi90"],
    "C": ["gol90", "ass90", "npxG90", "xA90", "tocchi90", "occ90", "tiri90"],
    "A": ["gol90", "ass90", "npxG90", "xA90", "tocchi90", "tiri90", "tiri_porta90"],
}

for ruolo in ("P", "D", "C", "A"):
    rows = data[ruolo]
    print(f"\n=== {ruolo} (n={len(rows)}) ===")
    for k in KEYS[ruolo]:
        vals = [x[k] for x in rows if x[k] > 0]
        if not vals:
            print(f"  {k}: nessun valore")
            continue
        print(f"  {k:14}  p50={pct(vals,50):.3f}  p75={pct(vals,75):.3f}  p90={pct(vals,90):.3f}  p95={pct(vals,95):.3f}  max={max(vals):.3f}")
