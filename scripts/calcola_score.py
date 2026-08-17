#!/usr/bin/env python3
"""Score 0-100 di convenienza: titolarita 35 + rating MV 25 + valore 20 + bonus reparto 20 - carte (max -5).

Principio: i minuti non vanno premiati due volte. La titolarita premia "quanto giochi";
il bonus reparto premia "quanto produci quando giochi", quindi usa statistiche /90
normalizzate sui percentili p90 del ruolo (scale FISSE, calcolate una tantum sul listone 26/27).
"""
import openpyxl

SRC = "/home/ubuntu/Downloads/euroleghe_master_classic_fotmob_2026_27_final_clean.xlsx"

# giornate per campionato (per la titolarita)
PARTITE = {"Liga": 38, "Premier League": 38, "Serie A": 38, "Bundesliga": 34, "Ligue 1": 34}

# riferimenti p90 per ruolo (calibrati sul listone, da analizza_percentili.py)
P90 = {
    "P": {"cs_rate": 0.50, "sub90": 1.8},
    "D": {"gol90": 0.181, "ass90": 0.246, "xA90": 0.156, "tocchi90": 2.423},
    "C": {"gol90": 0.351, "ass90": 0.367, "npxG90": 0.246, "xA90": 0.215, "tocchi90": 4.851, "tiri90": 2.489, "occ90": 1.959},
    "A": {"gol90": 0.721, "npxG90": 0.530, "tiri_porta90": 1.533, "tocchi90": 7.228},
}


def n(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def clamp(x, lo=0.0, hi=100.0):
    return max(lo, min(hi, x))


def norm(val, ref):
    return clamp(val / ref, 0.0, 1.0) if ref > 0 else 0.0


def bonus_reparto(ruolo, m, d):
    """m = minuti, d = dict dei campi /90. Ritorna 0..20."""
    if m <= 0:
        return 0.0
    r = P90[ruolo]
    if ruolo == "P":
        pres = d["presenze"]
        cs_rate = d["clean_sheet"] / pres if pres > 0 else 0.0
        sub90 = d["gol_subiti"] / m * 90
        return norm(cs_rate, r["cs_rate"]) * 10 + clamp((r["sub90"] - sub90) / r["sub90"], 0, 1) * 10
    if ruolo == "D":
        gol90 = d["gol"] / m * 90
        ass90 = d["assist"] / m * 90
        xA90 = d["xA"] / m * 90
        tocchi90 = d["tocchi_area"] / m * 90
        return (norm(gol90, r["gol90"]) * 4 + norm(ass90, r["ass90"]) * 4
                + norm(xA90, r["xA90"]) * 6 + norm(tocchi90, r["tocchi90"]) * 6)
    if ruolo == "C":
        gol90 = d["gol"] / m * 90
        ass90 = d["assist"] / m * 90
        npxG90 = d["npxG"] / m * 90
        xA90 = d["xA"] / m * 90
        tocchi90 = d["tocchi_area"] / m * 90
        tiri90 = d["tiri"] / m * 90
        occ90 = d["occ"] / m * 90
        return (norm(gol90, r["gol90"]) * 4 + norm(ass90, r["ass90"]) * 3
                + norm(npxG90, r["npxG90"]) * 3 + norm(xA90, r["xA90"]) * 3
                + norm(tocchi90, r["tocchi90"]) * 2 + norm(tiri90, r["tiri90"]) * 2
                + norm(occ90, r["occ90"]) * 3)
    if ruolo == "A":
        gol90 = d["gol"] / m * 90
        npxG90 = d["npxG"] / m * 90
        tiri_porta90 = d["tiri_in_porta"] / m * 90
        tocchi90 = d["tocchi_area"] / m * 90
        return (norm(gol90, r["gol90"]) * 8 + norm(npxG90, r["npxG90"]) * 6
                + norm(tiri_porta90, r["tiri_porta90"]) * 4 + norm(tocchi90, r["tocchi90"]) * 2)
    return 0.0


wb = openpyxl.load_workbook(SRC)
for sheet in ["Solo_Torneo", "Portieri", "Difensori", "Centrocampisti", "Attaccanti"]:
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    c = {v: i + 1 for i, v in enumerate(hdr) if v}
    if "score_finale" not in c:
        continue
    done = 0
    for r in range(2, ws.max_row + 1):
        ruolo = ws.cell(r, c["ruolo"]).value
        minuti = n(ws.cell(r, c["minuti_campionato"]).value)
        camp = ws.cell(r, c["campionato"]).value
        partite = PARTITE.get(camp, 38)
        tit = min(1.0, minuti / (partite * 90)) if minuti else 0.0
        if "titolarita_pct" in c:
            ws.cell(r, c["titolarita_pct"]).value = round(tit * 100, 1)

        rating = n(ws.cell(r, c["rating_fotmob"]).value)
        quot = n(ws.cell(r, c["quotazione"]).value)
        gialli = n(ws.cell(r, c["gialli"]).value)
        rossi = n(ws.cell(r, c["rossi"]).value)

        score = tit * 35
        score += clamp((rating - 6.0) / 2.0 * 25, 0, 25)
        score += clamp(20 - quot / 2.5, 0, 20)

        d = {
            "presenze": n(ws.cell(r, c["presenze_campionato"]).value),
            "clean_sheet": n(ws.cell(r, c["clean_sheet"]).value) if "clean_sheet" in c else 0,
            "gol_subiti": n(ws.cell(r, c["gol_subiti"]).value) if "gol_subiti" in c else 0,
            "gol": n(ws.cell(r, c["gol_campionato"]).value),
            "assist": n(ws.cell(r, c["assist_campionato"]).value),
            "npxG": n(ws.cell(r, c["xG_no_rigori"]).value),
            "xA": n(ws.cell(r, c["xA"]).value),
            "tocchi_area": n(ws.cell(r, c["tocchi_area"]).value),
            "tiri": n(ws.cell(r, c["tiri"]).value),
            "tiri_in_porta": n(ws.cell(r, c["tiri_in_porta"]).value),
            "occ": n(ws.cell(r, c["occasioni_create"]).value),
        }
        score += bonus_reparto(ruolo, minuti, d)
        score -= min(5.0, gialli * 0.5 + rossi * 3)

        ws.cell(r, c["score_finale"]).value = round(clamp(score, 0, 100))
        done += 1
    print(f"{sheet}: score calcolato per {done} giocatori")

wb.save(SRC)
print("salvato:", SRC)
