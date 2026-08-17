#!/usr/bin/env python3
"""Verifica se i 20 giocatori corretti dal fix_mismatch hanno presenze/minuti/titolarita non nulli."""
import openpyxl

MASTER = "/home/ubuntu/Downloads/euroleghe_master_classic_fotmob_2026_27_final_clean.xlsx"
FIX = ["Kim", "Gabriel Magalhaes", "Martinez Lis.", "Alvarez Y.", "Williams N.", "Williams I.",
       "Adeyemi", "Vivian", "Berenguer", "Ederson D.S.", "Garcia F.", "Navarro", "Ramos G.",
       "Cissè M.K.", "Gila", "Cresswell C.", "Joao Mario", "Chandler"]

wb = openpyxl.load_workbook(MASTER, data_only=True)
ws = wb["Solo_Torneo"]
h = [c.value for c in ws[1]]
c = {v: i + 1 for i, v in enumerate(h) if v}
print(f"{'nome':18} | pres | min | tit% | gol | ass")
for r in range(2, ws.max_row + 1):
    nome = ws.cell(r, c["giocatore"]).value
    if nome in FIX:
        pres = ws.cell(r, c["presenze_campionato"]).value
        minv = ws.cell(r, c["minuti_campionato"]).value
        tit = ws.cell(r, c["titolarita_pct"]).value
        gol = ws.cell(r, c["gol_campionato"]).value
        ass = ws.cell(r, c["assist_campionato"]).value
        flag = "  <-- PRESENZE NULL!" if pres is None else ""
        print(f"{nome:18} | {str(pres):4} | {str(minv):5} | {str(tit):4} | {str(gol):3} | {str(ass):3}{flag}")
