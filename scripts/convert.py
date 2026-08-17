#!/usr/bin/env python3
"""Converte Solo_Torneo + Piano_Rosa_26_500 -> players.json per l'app EuroFantaLeghe."""
import openpyxl, json, math

SRC = "/home/ubuntu/Downloads/euroleghe_master_classic_fotmob_2026_27_final_clean.xlsx"
OUT = "/home/ubuntu/apps/eurofantaleghe/public/data/players.json"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Solo_Torneo"]
h = [c.value for c in ws[1]]
c = {v: i + 1 for i, v in enumerate(h) if v}


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and math.isnan(v)) else v
    try:
        f = float(str(v).replace(",", "."))
        return None if math.isnan(f) else f
    except (ValueError, TypeError):
        return None


def txt(v):
    return None if v is None else str(v).strip()


# piano 26 -> nome -> offerta
plan = wb["Piano_Rosa_26_500"]
ph = [x.value for x in plan[1]]
pc = {v: i + 1 for i, v in enumerate(ph) if v}
offerte = {}
for r in range(2, plan.max_row + 1):
    nm = txt(plan.cell(r, pc["nome"]).value)
    off = num(plan.cell(r, pc["offerta_precisa"]).value)
    if nm and str(nm).strip():
        offerte[nm] = off

players = []
for r in range(2, ws.max_row + 1):
    ruolo = txt(ws.cell(r, c["ruolo"]).value)
    if ruolo not in ("P", "D", "C", "A"):
        continue
    nome = txt(ws.cell(r, c["giocatore"]).value)
    p = {
        "id": txt(ws.cell(r, c["id_fantacalcio"]).value),
        "nome": nome,
        "squadra": txt(ws.cell(r, c["squadra_fantacalcio"]).value),
        "campionato": txt(ws.cell(r, c["campionato"]).value),
        "ruolo": ruolo,
        "quotazione": num(ws.cell(r, c["quotazione"]).value),
        "base": num(ws.cell(r, c["prezzo_base_ceil"]).value),
        "fvm": num(ws.cell(r, c["fvm"]).value),
        "fotmob_id": num(ws.cell(r, c["fotmob_id"]).value),
        "eta": num(ws.cell(r, c["eta"]).value),
        "nazionalita": txt(ws.cell(r, c["nazionalita"]).value),
        "piede": txt(ws.cell(r, c["piede"]).value),
        "ruolo_reale": txt(ws.cell(r, c["ruolo_reale_principale"]).value),
        "presenze": num(ws.cell(r, c["presenze_campionato"]).value),
        "minuti": num(ws.cell(r, c["minuti_campionato"]).value),
        "titolarita_pct": num(ws.cell(r, c["titolarita_pct"]).value),
        "gol": num(ws.cell(r, c["gol_campionato"]).value),
        "assist": num(ws.cell(r, c["assist_campionato"]).value),
        "rating": num(ws.cell(r, c["rating_fotmob"]).value),
        "gialli": num(ws.cell(r, c["gialli"]).value),
        "rossi": num(ws.cell(r, c["rossi"]).value),
        "xG": num(ws.cell(r, c["xG"]).value),
        "xA": num(ws.cell(r, c["xA"]).value),
        "tiri": num(ws.cell(r, c["tiri"]).value),
        "tiri_in_porta": num(ws.cell(r, c["tiri_in_porta"]).value),
        "rigori_gol": num(ws.cell(r, c["rigori_gol"]).value),
        "xG_no_rigori": num(ws.cell(r, c["xG_no_rigori"]).value),
        "occasioni_create": num(ws.cell(r, c["occasioni_create"]).value),
        "big_chances": num(ws.cell(r, c["big_chances_create"]).value),
        "tocchi_area": num(ws.cell(r, c["tocchi_area"]).value),
        "tiri_punizione": num(ws.cell(r, c["tiri_punizione"]).value),
        "tiri_da_corner": num(ws.cell(r, c["tiri_da_corner"]).value),
        "crosses": num(ws.cell(r, c["crosses"]).value),
        "injury_status": txt(ws.cell(r, c["injury_status"]).value),
        "clean_sheet": num(ws.cell(r, c["clean_sheet"]).value),
        "gol_subiti": num(ws.cell(r, c["gol_subiti"]).value),
        "rigori_parati": num(ws.cell(r, c["rigori_parati"]).value),
        "score_finale": num(ws.cell(r, c["score_finale"]).value),
        "categoria": txt(ws.cell(r, c["categoria_auto"]).value),
        "max_bid": num(ws.cell(r, c["max_bid"]).value),
        "offerta": offerte.get(nome),
    }
    players.append(p)

import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(players, f, ensure_ascii=False)

# report
precomp = sum(1 for p in players if p["offerta"] is not None)
print("giocatori:", len(players))
print("precompilati con offerta:", precomp)
tot = sum((p["offerta"] or 0) for p in players)
print("totale offerte precompilate:", int(tot))
print("scritto:", OUT)
