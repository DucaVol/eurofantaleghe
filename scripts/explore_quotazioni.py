#!/usr/bin/env python3
"""Esplora la struttura delle quotazioni ufficiali EuroLeghe e del master per preparare la verifica."""
import openpyxl

QUOT = "/home/ubuntu/Downloads/Quotazioni_Fantacalcio_EuroLeghe_Stagione_2026_27.xlsx"
MASTER = "/home/ubuntu/Downloads/euroleghe_master_classic_fotmob_2026_27_final_clean.xlsx"

print("=" * 70)
print("QUOTAZIONI EuroLeghe 2026-27")
print("=" * 70)
wb = openpyxl.load_workbook(QUOT, data_only=True)
print("fogli:", wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n--- foglio '{sn}': {ws.max_row} righe x {ws.max_column} colonne ---")
    for r in range(1, min(4, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 15))]
        print(f"  r{r}: {vals}")

print()
print("=" * 70)
print("MASTER Solo_Torneo (colonne)")
print("=" * 70)
wb2 = openpyxl.load_workbook(MASTER, data_only=True)
ws = wb2["Solo_Torneo"]
hdr = [c.value for c in ws[1]]
print("numero colonne:", len([h for h in hdr if h]))
for i, h in enumerate(hdr):
    if h:
        print(f"  col{i+1:2}: {h}")
print("\nrecord campione (r2):")
for i, h in enumerate(hdr):
    if h:
        print(f"  {h}: {ws.cell(2, i+1).value}")
