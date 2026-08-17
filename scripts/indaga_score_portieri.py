#!/usr/bin/env python3
"""Indaga lo score dei portieri: top per score_finale con i parametri che lo compongono."""
import openpyxl

MASTER = "/home/ubuntu/Downloads/euroleghe_master_classic_fotmob_2026_27_final_clean.xlsx"
wb = openpyxl.load_workbook(MASTER, data_only=True)
ws = wb["Solo_Torneo"]
h = [c.value for c in ws[1]]
c = {v: i + 1 for i, v in enumerate(h) if v}

rows = []
for r in range(2, ws.max_row + 1):
    if ws.cell(r, c["ruolo"]).value != "P":
        continue
    nome = ws.cell(r, c["giocatore"]).value
    rows.append({
        "nome": nome,
        "sq": ws.cell(r, c["squadra_fantacalcio"]).value,
        "pres": ws.cell(r, c["presenze_campionato"]).value,
        "min": ws.cell(r, c["minuti_campionato"]).value,
        "tit": ws.cell(r, c["titolarita_pct"]).value,
        "rat": ws.cell(r, c["rating_fotmob"]).value,
        "cs": ws.cell(r, c["clean_sheet"]).value,
        "rp": ws.cell(r, c["rigori_parati"]).value,
        "base": ws.cell(r, c["prezzo_base"]).value,
        "quot": ws.cell(r, c["quotazione"]).value,
        "score": ws.cell(r, c["score_finale"]).value,
    })

rows.sort(key=lambda x: -(x["score"] or 0))
print(f"{'#':2} {'nome':16} {'sq':16} {'pres':>4} {'min':>5} {'tit%':>5} {'rat':>5} {'cs':>3} {'rp':>3} {'base':>4} {'quot':>4} {'score':>5}")
for i, p in enumerate(rows[:15], 1):
    print(f"{i:2} {p['nome']:16} {str(p['sq']):16} {str(p['pres']):>4} {str(p['min']):>5} "
          f"{str(p['tit']):>5} {str(p['rat']):>5} {str(p['cs']):>3} {str(p['rp']):>3} "
          f"{str(p['base']):>4} {str(p['quot']):>4} {str(p['score']):>5}")

print("\n--- Terracciano nel dettaglio ---")
for p in rows:
    if p["nome"] == "Terracciano":
        print(p)
