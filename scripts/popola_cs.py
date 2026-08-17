#!/usr/bin/env python3
"""Popola clean_sheet / gol_subiti / rigori_parati per i portieri nel Solo_Torneo e Portieri."""
import openpyxl

SRC = "/home/ubuntu/Downloads/euroleghe_master_classic_fotmob_2026_27_final_clean.xlsx"

# nome_file -> (clean_sheet, gol_subiti, rigori_parati, nota)
data = {
    "Raya": (19, 26, 0, ""),
    "Martinez E.": (7, 39, 1, ""),
    "Carnesecchi": (13, 35, 2, ""),
    "Unai Simon": (6, 54, 2, ""),
    "Oblak": (11, 31, 0, ""),
    "Garcia J.": (15, 21, 0, ""),
    "Flekken": (5, 38, 0, ""),
    "Neuer": (6, 20, 0, ""),
    "Valles": (8, 36, 1, ""),
    "Skorupski": (7, 15, 0, ""),
    "Kobel": (15, 34, 0, ""),
    "Petrovic D.": (11, 54, 0, ""),
    "Verbruggen": (10, 46, 1, ""),
    "Sanchez Ro.": (9, 47, 0, ""),
    "Butez": (19, 29, 1, ""),
    "Zetterer": (5, 33, 0, ""),
    "De Gea": (10, 49, 1, ""),
    "Provedel": (12, 27, 1, ""),
    "Di Gregorio": (13, 25, 1, ""),
    "Vandevoordt": (2, 15, 1, ""),
    "Alisson": (8, 31, 0, ""),
    "Donnarumma G.": (15, 29, 1, ""),
    "Lammens": (8, 39, 0, ""),
    "Maignan": (13, 35, 2, ""),
    "Kohn": (5, 27, 0, ""),
    "Milinkovic-Savic V.": (11, 24, 3, ""),
    "Jaouen": (15, 35, 1, "stats Ligue 2 25/26"),
    "Rulli": (8, 34, 0, "in uscita dal Marsiglia"),
    "Chevalier": (9, 13, 1, ""),
    "Courtois": (13, 28, 0, ""),
    "Samba": (10, 44, 2, ""),
    "Svilar": (18, 31, 0, ""),
    "Seimen": (9, 45, 0, "stats 2.Bundesliga + INF coscia"),
    "Dubravka": (4, 71, 0, ""),
    "Luiz Junior": (0, 2, 0, "solo 1 partita 26/27"),
    "Hornicek": (12, 33, 2, ""),
}

wb = openpyxl.load_workbook(SRC)
for sheet in ["Solo_Torneo", "Portieri"]:
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    col = {v: i + 1 for i, v in enumerate(hdr) if v}
    if "clean_sheet" not in col:
        # aggiunge 3 colonne in coda
        base = len(hdr) + 1
        ws.cell(1, base, "clean_sheet")
        ws.cell(1, base + 1, "gol_subiti")
        ws.cell(1, base + 2, "rigori_parati")
        ws.cell(1, base + 3, "cs_note")
        col["clean_sheet"] = base
        col["gol_subiti"] = base + 1
        col["rigori_parati"] = base + 2
        col["cs_note"] = base + 3
    found = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col["ruolo"]).value != "P":
            continue
        nome = ws.cell(r, col["giocatore"]).value
        if nome in data:
            cs, ga, ps, note = data[nome]
            ws.cell(r, col["clean_sheet"], cs)
            ws.cell(r, col["gol_subiti"], ga)
            ws.cell(r, col["rigori_parati"], ps)
            ws.cell(r, col["cs_note"], note)
            found += 1
    print(f"{sheet}: popolati {found} portieri su {len(data)}")

wb.save(SRC)
print("salvato:", SRC)

# verifica chi non è stato trovato
presenti = set()
ws = wb["Solo_Torneo"]
hdr = [c.value for c in ws[1]]
col = {v: i + 1 for i, v in enumerate(hdr) if v}
for r in range(2, ws.max_row + 1):
    if ws.cell(r, col["ruolo"]).value == "P":
        presenti.add(ws.cell(r, col["giocatore"]).value)
mancanti = set(data) - presenti
print("nomi nel dict non trovati nel file:", mancanti if mancanti else "nessuno")
