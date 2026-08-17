#!/usr/bin/env python3
"""Popola fantamedia (Fm) da Statistiche_Fantacalcio_EuroLeghe_2025_26 nel Solo_Torneo."""
import openpyxl

FANTA = "/home/ubuntu/Downloads/Statistiche_Fantacalcio_EuroLeghe_Stagione_2025_26.xlsx"
SRC = "/home/ubuntu/Downloads/euroleghe_master_classic_fotmob_2026_27_final_clean.xlsx"

# carica fantamedia: id -> (nome, fm, mv)
wb_f = openpyxl.load_workbook(FANTA, data_only=True)
ws_f = wb_f["Tutti"]
# header reale alla riga 2
hdr = [c.value for c in ws_f[2]]
idx = {v: i + 1 for i, v in enumerate(hdr) if v}
print("header fantamedia:", hdr)
fanta = {}
for r in range(3, ws_f.max_row + 1):
    fid = ws_f.cell(r, idx["Id"]).value
    fm = ws_f.cell(r, idx["Fm"]).value
    mv = ws_f.cell(r, idx["Mv"]).value
    nome = ws_f.cell(r, idx["Nome"]).value
    if fid is not None:
        fanta[int(fid)] = {"nome": nome, "fm": fm, "mv": mv}
print("fantamedia caricate:", len(fanta))


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


wb = openpyxl.load_workbook(SRC)
for sheet in ["Solo_Torneo", "Portieri", "Difensori", "Centrocampisti", "Attaccanti"]:
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    c = {v: i + 1 for i, v in enumerate(hdr) if v}
    if "fantamedia" not in c:
        base = len(hdr) + 1
        ws.cell(1, base, "fantamedia")
        c["fantamedia"] = base
    found = 0
    tot = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, c["ruolo"]).value not in ("P", "D", "C", "A"):
            continue
        tot += 1
        fid = ws.cell(r, c["id_fantacalcio"]).value
        if fid is not None and int(fid) in fanta:
            ws.cell(r, c["fantamedia"], num(fanta[int(fid)]["fm"]))
            found += 1
    print(f"{sheet}: fantamedia popolata per {found}/{tot}")

wb.save(SRC)
print("salvato:", SRC)
