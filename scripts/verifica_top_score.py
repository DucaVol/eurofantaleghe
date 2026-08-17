#!/usr/bin/env python3
"""Stampa i top per ruolo secondo il nuovo score, per verificare che la classifica sia sensata."""
import openpyxl

MASTER = "/home/ubuntu/Downloads/euroleghe_master_classic_fotmob_2026_27_final_clean.xlsx"
wb = openpyxl.load_workbook(MASTER, data_only=True)
ws = wb["Solo_Torneo"]
h = [c.value for c in ws[1]]
c = {v: i + 1 for i, v in enumerate(h) if v}

rows = {r: [] for r in ("P", "D", "C", "A")}
for r in range(2, ws.max_row + 1):
    ruolo = ws.cell(r, c["ruolo"]).value
    if ruolo not in rows:
        continue
    rows[ruolo].append({
        "nome": ws.cell(r, c["giocatore"]).value,
        "sq": ws.cell(r, c["squadra_fantacalcio"]).value,
        "min": ws.cell(r, c["minuti_campionato"]).value,
        "tit": ws.cell(r, c["titolarita_pct"]).value,
        "rat": ws.cell(r, c["rating_fotmob"]).value,
        "quot": ws.cell(r, c["quotazione"]).value,
        "gol": ws.cell(r, c["gol_campionato"]).value,
        "ass": ws.cell(r, c["assist_campionato"]).value,
        "score": ws.cell(r, c["score_finale"]).value,
    })

for ruolo in ("P", "D", "C", "A"):
    rows[ruolo].sort(key=lambda x: -(x["score"] or 0))
    print(f"\n=== {ruolo} — top 10 ===")
    print(f"{'nome':18}{'sq':16}{'min':>5}{'tit%':>6}{'rat':>5}{'quot':>5}{'gol':>4}{'ass':>4}{'score':>6}")
    for p in rows[ruolo][:10]:
        print(f"{p['nome']:18}{str(p['sq']):16}{str(p['min']):>5}{str(p['tit']):>6}{str(p['rat']):>5}"
              f"{str(p['quot']):>5}{str(p['gol']):>4}{str(p['ass']):>4}{str(p['score']):>6}")
